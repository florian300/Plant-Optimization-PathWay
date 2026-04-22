import pandas as pd
import numpy as np
import os
import shutil
from rich import print
from tqdm import tqdm
from .optimizer import PathFinderOptimizer
import plotly.graph_objects as go
import plotly.io as pio
import plotly.utils
import json
from .plots.carbon_tax import build_carbon_tax_figure
from pathway.core.plots.financial import (
    build_transition_cost_figure, 
    build_external_financing_figure,
    build_interest_paid_figure,
    build_mac_figure
)
from pathway.core.plots.carbon import (
    build_carbon_price_figure, 
    build_co2_trajectory_figure,
    build_indirect_emissions_figure
)
from pathway.core.plots.energy_mix import build_resources_mix_figure
from pathway.core.plots.investment import build_investment_plan_figure
from pathway.core.plots.opex import build_opex_figure
from pathway.core.plots.prices import build_simulation_prices_figure

class PathFinderReporter:
    def __init__(self, optimizer: PathFinderOptimizer, scenario_id: str = "DEFAULT", scenario_name: str = "Default", generate_excel: bool = True, verbose: bool = False, progress_cb=None):
        self.opt = optimizer
        self.data = optimizer.data
        self.years = optimizer.years
        self.scenario_id = scenario_id
        self.scenario_name = scenario_name
        self.generate_excel = generate_excel
        self.verbose = verbose
        self.progress_cb = progress_cb
        repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
        self.entity_id = getattr(self.opt.entity, 'id', 'Base Company') if hasattr(self, 'opt') and hasattr(self.opt, 'entity') else 'Base Company'
        self.entity_name = getattr(self.opt.entity, 'name', self.entity_id) if hasattr(self, 'opt') and hasattr(self.opt, 'entity') else 'Base Company'
        
        safe_name = "".join([c for c in self.entity_name if c.isalpha() or c.isdigit() or c in ' -_']).strip()
        if not safe_name: safe_name = self.entity_id
        
        self.results_dir = os.path.join(repo_root, 'artifacts', 'reports', safe_name, self.scenario_name)
        self.charts_data = [] # List of (title, dataframe) tuples

    def _save_plotly_figure(self, fig, base_filename, show_png: bool = True):
        """Unified save method for Plotly figures (PNG + JSON)."""
        charts_dir = os.path.join(self.results_dir, "charts")
        os.makedirs(charts_dir, exist_ok=True)
        
        # 1. Export static PNG
        # ONLY if show_png is enabled.
        if show_png:
            png_path = os.path.join(self.results_dir, f"{self.scenario_name}_{base_filename}.png")
            fig.write_image(png_path, scale=2, width=1200, height=800)
        
        # 2. Export JSON for web dashboard (ALWAYS)
        json_path = os.path.join(charts_dir, f"{base_filename}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            f.write(fig.to_json())

    def _save_plotly_placeholder(self, base_filename: str, title: str, message: str, show_png: bool = True):
        """Creates a styled Plotly placeholder when no data is available, keeping artifact parity."""
        fig = go.Figure()
        
        # Add a central annotation with the message
        fig.add_annotation(
            text=f"<b>{title}</b><br><br>{message}",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=16, color="#64748b"),
            align="center"
        )
        
        # Simple styled layout
        fig.update_layout(
            template="plotly_white",
            xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
            yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
            margin=dict(l=40, r=40, t=60, b=40),
            height=600
        )
        
        self._save_plotly_figure(fig, base_filename, show_png=show_png)

    def _primary_emission_resource(self) -> str:
        return self.opt.entity.primary_emission_resource() or 'CO2_EM'

    def _is_primary_emission_resource(self, resource_id: str) -> bool:
        primary = self._primary_emission_resource()
        # Case-insensitive check and fallback if primary contains CO2
        return resource_id == primary or (primary == 'CO2_EM' and 'CO2' in str(resource_id).upper())

    def _process_emission_baseline(self, process) -> float:
        return self.opt.entity.process_emission_baseline(process, self._primary_emission_resource())

    def _process_primary_energy_consumption(self, process) -> float:
        # Filter to only energy resources for MW-based scaling
        excluded = self._get_non_energy_resource_ids()
        val, _ = process.primary_energy_consumption(self.opt.entity.base_consumptions, excluded_resources=excluded)
        return val

    def _tech_emission_impact(self, tech):
        primary = self._primary_emission_resource()
        imp = tech.impacts.get(primary)
        if not imp and primary == 'CO2_EM':
            co2_keys = [k for k in tech.impacts.keys() if 'CO2' in k]
            if co2_keys:
                imp = tech.impacts[co2_keys[0]]
        return imp

    def _get_energy_resource_ids(self) -> set:
        """Returns IDs of resources that are considered 'Energy' for scaling purposes."""
        energy_ids = set()
        for res_id, res in self.data.resources.items():
            if (res_id.startswith('EN_') or 
                res.category.upper() in ['ENERGY', 'FUEL'] or 
                res.resource_type.upper() in ['ELECTRICITY', 'FUEL', 'GAS', 'HYDROGEN']):
                energy_ids.add(res_id)
        return energy_ids

    def _get_non_energy_resource_ids(self) -> set:
        """Returns IDs of resources that should be EXCLUDED from scaling drivers (e.g. Water, CO2)."""
        all_ids = set(self.data.resources.keys())
        energy_ids = self._get_energy_resource_ids()
        return all_ids - energy_ids

    def _get_safe_var_value(self, var, default=0.0) -> float:
        """Returns varValue sanitized for None or extreme junk values on infeasibility."""
        if var is None: return default
        val = getattr(var, 'varValue', None)
        if val is None: return default
        # Junk data from CBC on infeasibility is often ~1e10
        if abs(val) > 1e11: 
            return default
        return float(val)

    # Mathematical parity: tCO2-based scaling remains "entity emission baseline x process share"
    # and MW-based scaling remains "dominant weighted energy driver / operating hours".
        
    def generate_report(self):
        # 0. Calculate total steps for progress tracking
        toggles = self.data.reporting_toggles
        # Count active charts
        steps_total = sum(1 for k, v in vars(toggles).items() if k.startswith('chart_') and v is True)
        # Add Excel export steps (Data collection + Save)
        steps_total += 2 if self.generate_excel else 0
        steps_done = 0

        def _step():
            nonlocal steps_done
            steps_done += 1
            if self.progress_cb:
                self.progress_cb(steps_done, steps_total)

        if self.verbose:
            print(f"  [yellow][Reporter][/yellow] [STATS] Generating reports ({steps_total} steps)...")
        
        # 0. Clear and Recreate Results Directory
        results_dir = self.results_dir
        if os.path.exists(results_dir):
            try:
                shutil.rmtree(results_dir)
            except Exception as e:
                print(f"  [yellow][Reporter][/yellow] [!] Could not clear results directory: {e}")
        os.makedirs(results_dir, exist_ok=True)

        # 1. Collect Data
        if self.verbose:
            print("  [yellow][Reporter][/yellow] [SEARCH] Collecting results data...")
        _step() # Mark data collection step
        # ... (rest of data collection remains same)
        investments = []
        for t in tqdm(self.years, desc="Processing Years", leave=False, disable=not self.verbose):
            for p_id, process in self.opt.entity.processes.items():
                for t_id in process.valid_technologies:
                    tech = self.data.technologies[t_id]
                    if tech.is_continuous_improvement:
                        continue
                    invest_val = self._get_safe_var_value(self.opt.invest_vars[(t, p_id, t_id)])
                    if invest_val > 1e-6:
                        invested_units = invest_val
                        
                        # Calculate CAPEX scaled by fractional units (tCO2 or MW)
                        cap_capex = 1.0
                        if tech.capex_per_unit:
                            if tech.capex_unit == 'tCO2': 
                                cap_capex = self.opt.entity.process_emission_baseline(process, self._primary_emission_resource())
                            elif 'MW' in tech.capex_unit.upper(): 
                                best_val = self._process_primary_energy_consumption(process)
                                cap_capex = best_val / self.opt.entity.annual_operating_hours
                        current_capex = tech.capex_by_year.get(t, tech.capex)
                        capex_cost = (current_capex * cap_capex / process.nb_units) * invested_units
                        
                        impact_strings = []
                        primary_emis = self._primary_emission_resource()
                        for res_id, imp in tech.impacts.items():
                            val = imp['value']
                            if self._is_primary_emission_resource(res_id):
                                base_c = self.opt.entity.process_emission_baseline(process, primary_emis)
                            else:
                                base_c = self.opt.entity.base_consumptions.get(res_id, 0.0) * process.consumption_shares.get(res_id, 0.0)
                                
                            target_change = 0.0
                            if imp['type'] == 'variation' or imp['type'] == 'up':
                                target_change = val * base_c
                            elif imp['type'] == 'new':
                                ref_res = imp.get('ref_resource')
                                if self._is_primary_emission_resource(ref_res):
                                    base_ref_amount = self.opt.entity.process_emission_baseline(process, primary_emis)
                                elif ref_res and ref_res in self.opt.entity.base_consumptions:
                                    base_ref_amount = self.opt.entity.base_consumptions.get(ref_res, 0.0) * process.consumption_shares.get(ref_res, 0.0)
                                else:
                                    base_ref_amount = 1.0
                                    
                                if imp.get('reference', '') == 'AVOIDED' and ref_res:
                                    ref_imp = tech.impacts.get(ref_res)
                                    if not ref_imp and self._is_primary_emission_resource(ref_res):
                                        imp_keys = [k for k in tech.impacts.keys() if self._is_primary_emission_resource(k)]
                                        if imp_keys: ref_imp = tech.impacts[imp_keys[0]]
                                    if ref_imp and ref_imp['type'] in ['variation', 'up'] and ref_imp['value'] < 0:
                                        base_ref_amount = abs(ref_imp['value']) * base_ref_amount

                                # Fallback just in case for PEM_H2 original logic if no ref_resource
                                if res_id == 'EN_ELEC' and t_id == 'PEM_H2' and not ref_res:
                                    # Fallback to fuel driver as reference
                                    ref_val, _ = process.primary_energy_consumption(self.opt.entity.base_consumptions)
                                    base_ref_amount = ref_val * 0.5
                                    
                                target_change = val * base_ref_amount
                                
                            # Scale the targeted baseline impact by the number of active units being invested
                            target_change = target_change * (invested_units / process.nb_units)
                                
                            if abs(target_change) > 1e-4:
                                direction = "Decrease" if target_change < 0 else "Increase"
                                impact_strings.append(f"{direction} {abs(target_change):,.2f} {res_id}")
                        
                        investments.append({
                            'Year': t, 
                            'Process': p_id, 
                            'Technology': t_id, 
                            'Units_Invested': f"{invested_units:.2f}/{process.nb_units}" if invested_units % 1 != 0 else f"{int(invested_units)}/{process.nb_units}",
                            'Capex_Euros': capex_cost,
                            'Impacts': " | ".join(impact_strings)
                        })
                    
        df_invest = pd.DataFrame(investments)
        
        # Consumptions
        cons_data = []
        for t in self.years:
            row = {'Year': t}
            primary_emis = self._primary_emission_resource()
            for res_id in self.data.resources:
                if not self._is_primary_emission_resource(res_id):
                    val = self.opt.cons_vars[(t, res_id)].varValue
                    res = self.data.resources[res_id]
                    if res.unit.upper() == 'GJ':
                        val = val / 3.6 # GJ to MWh
                    row[res_id] = val
            
            # --- Granular H2: Explicitly capture PRODUCED_ON_SITE ---
            if hasattr(self.opt, 'h2_supply_vars') and t in self.opt.h2_supply_vars:
                row['EN_H2_ON_SITE'] = self.opt.h2_supply_vars[t].get('PRODUCED_ON_SITE', 0.0).varValue if hasattr(self.opt.h2_supply_vars[t].get('PRODUCED_ON_SITE', 0.0), 'varValue') else 0.0
            
            cons_data.append(row)
        df_cons = pd.DataFrame(cons_data)
        
        # --- PRE-CALCULATE STRIKE PRICES FOR EXCEL ---
        # Map Year -> Max Strike Price
        excel_strikes = {t: 0.0 for t in self.years}
        if hasattr(self.opt, "ccfd_used_vars"):
            for (t_inv, p_id, t_id), var in self.opt.ccfd_used_vars.items():
                if var.varValue and var.varValue > 0.5:
                    tech = self.data.technologies[t_id]
                    base_p = self.data.time_series.carbon_prices.get(t_inv, 0.0)
                    strike_val = (1.0 + self.data.ccfd_params.eua_price_pct) * base_p
                    start_yr = t_inv + tech.implementation_time
                    end_yr = start_yr + self.data.ccfd_params.duration
                    for y in range(start_yr, end_yr):
                        if y in excel_strikes:
                            excel_strikes[y] = max(excel_strikes[y], strike_val)

        # Emissions
        emis_data = []
        for t in self.years:
            # Emis_vars is now strictly Direct Emissions
            direct_co2 = self.opt.emis_vars[t].varValue
            
            # Index for UP factor calculation
            yr_idx = list(self.opt.years).index(t)
            
            # Calculate Indirect Emissions for this year
            indirect_co2 = 0.0
            for r_id in self.data.resources:
                if r_id in self.data.time_series.other_emissions_factors:
                    factor = self.data.time_series.other_emissions_factors[r_id].get(t, 0.0)
                    if factor > 0:
                        cons_val = self._get_safe_var_value(self.opt.cons_vars.get((t, r_id)))
                        indirect_co2 += cons_val * factor
                        
            total_co2 = direct_co2 + indirect_co2
            
            if self.opt.entity.sv_act_mode == "PI":
                fq_pct = self.data.time_series.carbon_quotas_pi.get(t, 0.0)
            else:
                fq_pct = self.data.time_series.carbon_quotas_norm.get(t, 0.0)
                
            taxed_co2 = self.opt.taxed_emis_vars[t].varValue
            tax_price = self.data.time_series.carbon_prices.get(t, 0.0)
            penalty_factor = self.data.time_series.carbon_penalties.get(t, 0.0)
            
            # --- INDIRECT CARBON TAX (User Request) ---
            indirect_tax_cost = 0.0
            for r_id, res in self.data.resources.items():
                if res.tax_indirect_emissions:
                    factor = self.data.time_series.other_emissions_factors.get(r_id, {}).get(t, 0.0)
                    if factor > 0:
                        cons_val = self._get_safe_var_value(self.opt.cons_vars.get((t, r_id)))
                        indirect_tax_cost += (cons_val * factor * tax_price) / 1_000_000.0
            # Total cost is the sum of paid and penalized quotas
            # Differentiate standard tax from additional penalty
            paid_q = self.opt.paid_quota_vars[t].varValue or 0.0
            penal_q = self.opt.penalty_quota_vars[t].varValue or 0.0
            
            # The penalty list should contain the actual financial impact computed on penalty variables.
            # Tax cost covers the base tax rate for all quotas (both paid and penal portions).
            standard_tax_cost = (paid_q * tax_price + penal_q * tax_price) / 1_000_000.0
            quota_penalty_cost = (penal_q * tax_price * penalty_factor) / 1_000_000.0
            
            # --- AGGREGATE OBJECTIVE PENALTIES (Realistic only) ---
            # IMPORTANT: We only count objective penalties if they are NOT related to the primary CO2 
            # to avoid double counting the actual Carbon Tax in the "Tax" chart.
            obj_penalty_cost = 0.0
            for idx, obj in enumerate(self.data.objectives):
                if obj.penalty_type == "PENALTIES":
                    # Check for year-specific penalties (idx, t)
                    if (idx, t) in self.opt.penalty_vars:
                        v_val = self.opt.penalty_vars[(idx, t)].varValue or 0.0
                        if v_val > 1e-4:
                            # Only add if it's not the primary emission (which is already in standard_tax_cost)
                            if not self.opt._is_primary_emission_objective(obj.resource):
                                obj_penalty_cost += (v_val * tax_price * (1.0 + penalty_factor)) / 1_000_000.0
                    
                    # Check for global penalties
                    if idx in self.opt.penalty_vars and obj.target_year == t:
                        v_val = self.opt.penalty_vars[idx].varValue or 0.0
                        if v_val > 1e-4:
                            if not self.opt._is_primary_objective_resource(obj.resource):
                                obj_penalty_cost += (v_val * tax_price * (1.0 + penalty_factor)) / 1_000_000.0
            
            total_penalty_cost = quota_penalty_cost + obj_penalty_cost
            tax_cost_meuros = standard_tax_cost + total_penalty_cost
            
            # Calcul des émissions évitées (par rapport aux émissions de référence directes)
            avoided_total_kt = max(0.0, (self.opt.entity.base_emissions - direct_co2) / 1000.0)
            
            # Identify CCS/CCU contribution
            captured_kt = 0.0
            for p_id, process in self.opt.entity.processes.items():
                for t_id in process.valid_technologies:
                    tech = self.data.technologies[t_id]
                    if tech.is_continuous_improvement:
                        continue
                    act_var = self._get_safe_var_value(self.opt.active_vars[(t, p_id, t_id)])
                    if act_var > 1e-6:
                        if tech.tech_category == 'Carbon Capture':
                            primary_emis = self._primary_emission_resource()
                            imp = tech.impacts.get(primary_emis)
                            if not imp:
                                co2_keys = [k for k in tech.impacts.keys() if self._is_primary_emission_resource(k)]
                                if co2_keys: imp = tech.impacts[co2_keys[0]]
                            
                            if imp:
                                initial_val = self.opt.entity.process_emission_baseline(process, primary_emis)
                                ref_res = imp.get('ref_resource')
                                if self._is_primary_emission_resource(ref_res):
                                    base_ref_amount = initial_val
                                elif ref_res and ref_res in self.opt.entity.base_consumptions:
                                    base_ref_amount = self.opt.entity.base_consumptions.get(ref_res, 0.0) * process.consumption_shares.get(ref_res, 0.0)
                                else:
                                    base_ref_amount = 1.0
                                    
                                if imp.get('reference', '') == 'AVOIDED' and ref_res:
                                    ref_imp = tech.impacts.get(ref_res)
                                    if not ref_imp and self._is_primary_emission_resource(ref_res):
                                        imp_keys = [k for k in tech.impacts.keys() if self._is_primary_emission_resource(k)]
                                        if imp_keys: ref_imp = tech.impacts[imp_keys[0]]
                                    if ref_imp and ref_imp['type'] in ['variation', 'up'] and ref_imp['value'] < 0:
                                        base_ref = abs(ref_imp['value']) * base_ref_amount
                                    else:
                                        base_ref = base_ref_amount
                                else:
                                    base_ref = base_ref_amount

                                if imp['type'] == 'variation' or imp['type'] == 'up':
                                    reduction = -imp['value'] * initial_val
                                elif imp['type'] == 'new':
                                    reduction = -imp['value'] * base_ref
                                else:
                                    reduction = 0
                                    
                                # Scale reduction by active units proportion
                                reduction = reduction * (act_var / process.nb_units)
                                
                                # Apply UP compounding to the reduction if UP exists for this process
                                up_rate = 0.0
                                if 'UP' in process.valid_technologies:
                                    up_tech = self.data.technologies['UP']
                                    up_imp = up_tech.impacts.get('ALL') or up_tech.impacts.get(primary_emis)
                                    if up_imp and (up_imp['type'] == 'variation' or up_imp['type'] == 'up'):
                                        up_rate = abs(up_imp['value'])
                                reduction *= ((1.0 - up_rate) ** yr_idx)
                                
                                captured_kt += max(0.0, reduction / 1000.0)

            really_avoided_kt = max(0.0, avoided_total_kt - captured_kt)
            
            # DAC and Credits
            dac_cap = 0.0
            dac_limit_kt = 0.0
            if self.data.dac_params.active:
                dac_v = getattr(self.opt.dac_captured_vars.get(t), 'varValue', 0.0) or 0.0
                dac_cap = dac_v / 1000.0
                if self.data.dac_params.max_volume_pct < 1.0:
                    # Consistent with optimizer logic
                    ref_yr = self.data.dac_params.ref_year
                    # Prioritize historical/external reference from Entities sheet (nested dict)
                    yr_map = self.opt.entity.ref_baselines.get(self._primary_emission_resource(), {})
                    if ref_yr in yr_map:
                        ref_emis = yr_map[ref_yr]
                    elif yr_map:
                        ref_emis = next(iter(yr_map.values()))
                    else:
                        ref_emis = self.opt.entity.base_emissions

                    # For simulated years, look in emis_data
                    for d in emis_data:
                        if d['Year'] == ref_yr:
                            ref_emis = d['Direct_CO2']
                            break
                    dac_limit_kt = (ref_emis * self.data.dac_params.max_volume_pct) / 1000.0

            credit_vol = 0.0
            if self.data.credit_params.active:
                cred_v = getattr(self.opt.credit_purchased_vars.get(t), 'varValue', 0.0) or 0.0
                credit_vol = cred_v / 1000.0
                
            emis_data.append({
                'Year': t,
                'Direct_CO2': direct_co2,
                'Indirect_CO2': indirect_co2,
                'Total_CO2': total_co2,
                'DAC_Captured_kt': dac_cap,
                'DAC_Limit_kt': dac_limit_kt,
                'Credits_Purchased_kt': credit_vol,
                'Taxed_CO2': taxed_co2,
                'Free_Quota': direct_co2 * fq_pct if fq_pct <= 1.0 else fq_pct,  # Based on actual yearly emissions
                'Tax_Price': tax_price,
                'Standard_Tax_Cost_MEuros': standard_tax_cost,
                'Indirect_Tax_Cost_MEuros': indirect_tax_cost,
                'Penalty_Cost_MEuros': total_penalty_cost,
                'Tax_Cost_MEuros': tax_cost_meuros,
                'Avoided_Direct_CO2_kt': avoided_total_kt,
                'Avoided_Total_CO2_kt': max(0.0, (self.opt.entity.base_emissions + self.opt.entity.base_emissions * 0.1 - total_co2) / 1000.0), # Heuristic
                'Captured_CO2_kt': captured_kt,
                'Really_Avoided_CO2_kt': really_avoided_kt,
                'DAC_Captured_kt': dac_cap,
                'Credits_Purchased_kt': credit_vol,
                'Penalty_Factor': self.data.time_series.carbon_penalties.get(t, 0.0),
                'Effective_Price': tax_price * (1.0 + self.data.time_series.carbon_penalties.get(t, 0.0)),
                'Strike_Price': excel_strikes.get(t, 0.0)
            })
            
        df_emis = pd.DataFrame(emis_data)
        df_emis['Cumul_Tax_MEuros'] = df_emis['Tax_Cost_MEuros'].cumsum()

        # Generate Indirect Emissions Breakdown
        indir_breakdown = []
        for t in self.years:
            row = {'Year': t}
            for r_id in self.data.resources:
                if r_id in self.data.time_series.other_emissions_factors:
                    factor = self.data.time_series.other_emissions_factors[r_id].get(t, 0.0)
                    if factor > 0:
                        cons_val = self.opt.cons_vars.get((t, r_id))
                        if cons_val is not None:
                            val = self._get_safe_var_value(cons_val)
                            if r_id == 'EN_ELEC':
                                base_elec = self.opt.entity.base_consumptions.get('EN_ELEC', 0.0)
                                new_elec = max(0.0, val - base_elec)
                                row['EN_ELEC_BASE'] = base_elec * factor
                                row['EN_ELEC_TECH'] = new_elec * factor
                            else:
                                row[r_id] = val * factor
            indir_breakdown.append(row)
        df_indir = pd.DataFrame(indir_breakdown).fillna(0.0)
        
        # 1.5 Technology Costs (CAPEX & OPEX)
        # We use a two-pass approach to distribute aids correctly.
        aids_dist = {} # (year, col_name) -> value
        tech_capex_spent = {} # (year, p_id, t_id) -> value
        tech_co2_abatement = {} # (year, p_id, t_id) -> value in tCO2 abated
        tech_invested_procs = {} # (year, p_id, t_id) -> list of process names (for labels)
        new_investments = set() # (year, p_id, t_id) -> tracking for arrows
        
        for t in self.years:
            for t_id in self.data.technologies:
                aids_dist[(t, f"Aid_GRANT_{t_id}")] = 0.0
                aids_dist[(t, f"Aid_CCFD_{t_id}")] = 0.0
            for p_id in self.opt.entity.processes:
                for t_id in self.data.technologies:
                    tech_capex_spent[(t, p_id, t_id)] = 0.0
                    tech_co2_abatement[(t, p_id, t_id)] = 0.0
                    tech_invested_procs[(t, p_id, t_id)] = []

        for t in self.years:
            for t_id, tech in self.data.technologies.items():
                if tech.is_continuous_improvement:
                    continue
                for p_id, process in self.opt.entity.processes.items():
                    if t_id in process.valid_technologies:
                        # CAPEX calculation (scaled by tCO2 or MW)
                        cap_capex = 1.0
                        if tech.capex_per_unit:
                            if tech.capex_unit == 'tCO2': 
                                cap_capex = self.opt.entity.process_emission_baseline(process, self._primary_emission_resource())
                            elif 'MW' in tech.capex_unit.upper(): 
                                best_val = self._process_primary_energy_consumption(process)
                                cap_capex = best_val / self.opt.entity.annual_operating_hours
                        
                        invest_v = self._get_safe_var_value(self.opt.invest_vars[(t, p_id, t_id)])
                        if invest_v > 1e-6:
                            current_capex = tech.capex_by_year.get(t, tech.capex)
                            true_capex = (current_capex * cap_capex / process.nb_units) * invest_v
                            if true_capex > 0 and self.verbose:
                                print(f"  [yellow][Reporter][/yellow] [DEBUG] {t}: {p_id} {t_id} invest={invest_v} capex={current_capex} true_capex={true_capex}")
                            tech_capex_spent[(t, p_id, t_id)] += true_capex
                            
                            # Calculate CO2 Abatement added by this investment
                            primary_emis = self._primary_emission_resource()
                            imp = tech.impacts.get(primary_emis)
                            if not imp:
                                imp_keys = [k for k in tech.impacts.keys() if self._is_primary_emission_resource(k)]
                                if imp_keys: imp = tech.impacts[imp_keys[0]]
                            
                            if imp:
                                initial_val = self.opt.entity.process_emission_baseline(process, primary_emis)
                                yr_idx = list(self.opt.years).index(t)
                                
                                # UP compounding for the baseline at time of investment
                                up_rate = 0.0
                                if 'UP' in process.valid_technologies:
                                    up_tech = self.data.technologies['UP']
                                    up_imp = up_tech.impacts.get('ALL') or up_tech.impacts.get(primary_emis)
                                    if up_imp and (up_imp['type'] == 'variation' or up_imp['type'] == 'up'):
                                        up_rate = abs(up_imp['value'])
                                initial_val *= ((1.0 - up_rate) ** yr_idx)

                                ref_res = imp.get('ref_resource')
                                if self._is_primary_emission_resource(ref_res):
                                    base_ref = initial_val
                                elif ref_res and ref_res in self.opt.entity.base_consumptions:
                                    base_ref = self.opt.entity.base_consumptions.get(ref_res, 0.0) * process.consumption_shares.get(ref_res, 0.0)
                                else:
                                    # Fallback to initial_val if ref_res not found or same as CO2
                                    base_ref = initial_val if initial_val > 0 else 1.0

                                if imp['type'] == 'variation' or imp['type'] == 'up':
                                    red = -imp['value'] * initial_val
                                elif imp['type'] == 'new':
                                    red = -imp['value'] * base_ref
                                else:
                                    red = 0
                                
                                tech_co2_abatement[(t, p_id, t_id)] += red * (invest_v / process.nb_units)
                            
                            new_investments.add((t, p_id, t_id)) # Mark as NEW for arrows
                            
                            proc_display = f"({invest_v:.2f}/{process.nb_units})" if invest_v % 1 != 0 else f"({int(invested_units)}/{process.nb_units})"
                            tech_invested_procs[(t, p_id, t_id)].append(proc_display)
                            
                            # GRANT: Use precise value from the lp variable
                            if hasattr(self.opt, 'grant_amt_vars') and self.opt.grant_amt_vars.get((t, p_id, t_id)):
                                grant_val = self._get_safe_var_value(self.opt.grant_amt_vars[(t, p_id, t_id)])
                                aids_dist[(t, f"Aid_GRANT_{t_id}")] += grant_val
                                    
                            # CCfD: Distributed over contract duration
                            if hasattr(self.opt, 'ccfd_used_vars') and self.opt.ccfd_used_vars.get((t, p_id, t_id)):
                                if self._get_safe_var_value(self.opt.ccfd_used_vars[(t, p_id, t_id)]) > 0.5:
                                    ccfd_p = self.data.ccfd_params
                                    if not imp:
                                        imp_keys = [k for k in tech.impacts.keys() if self._is_primary_emission_resource(k)]
                                        if imp_keys: imp = tech.impacts[imp_keys[0]]
                                        
                                    if imp and (imp['type'] == 'variation' or imp['type'] == 'up'):
                                        reduction_frac = (-imp['value'] if imp['value'] < 0 else 0) / process.nb_units
                                        if reduction_frac > 0:
                                            max_emis = self.opt.entity.process_emission_baseline(process, primary_emis)
                                            # Scale avoided tons by units invested
                                            avoided_tons = reduction_frac * max_emis * invest_v
                                            start_yr = t + tech.implementation_time
                                            end_yr = start_yr + ccfd_p.duration
                                            
                                            for tau in range(start_yr, end_yr):
                                                if tau in self.years:
                                                    c_price = self.data.time_series.carbon_prices.get(tau, 0.0)
                                                    strike_price = (1.0 + ccfd_p.eua_price_pct) * self.data.time_series.carbon_prices.get(t, 0.0)
                                                    if ccfd_p.contract_type == 2:
                                                        subsidy = strike_price - c_price
                                                    else:
                                                        subsidy = max(0, strike_price - c_price)
                                                    aids_dist[(tau, f"Aid_CCFD_{t_id}")] += (subsidy * avoided_tons)

            if self.data.dac_params.active:
                dac_added_v = self._get_safe_var_value(self.opt.dac_added_capacity_vars.get(t))
                if dac_added_v > 1e-4:
                    dac_capex = dac_added_v * self.data.dac_params.capex_by_year.get(t, 0.0)
                    tech_capex_spent[(t, 'INDIRECT', 'DAC')] = dac_capex
                    tech_co2_abatement[(t, 'INDIRECT', 'DAC')] = dac_added_v
                    new_investments.add((t, 'INDIRECT', 'DAC'))
                    tech_invested_procs[(t, 'INDIRECT', 'DAC')] = [f"{dac_added_v / 1000:,.0f}ktCO2"]
            
            if self.data.credit_params.active:
                cred_v = self._get_safe_var_value(self.opt.credit_purchased_vars.get(t))
                if cred_v > 1e-4:
                    tech_capex_spent[(t, 'INDIRECT', 'CREDIT')] = 0.0 # No CAPEX for credits
                    tech_invested_procs[(t, 'INDIRECT', 'CREDIT')] = [f"{cred_v / 1000:,.0f}ktCO2"]

        # Identify all relevant (p_id, t_id) projects for granular dataset
        project_ids = sorted(list({(p_id, t_id) for (y, p_id, t_id), v in tech_capex_spent.items() if v > 1e-9}))
        if self.verbose:
            print(f"  [yellow][Reporter][/yellow] [DEBUG] Found {len(project_ids)} projects with non-zero investment.")
            if len(project_ids) > 0:
                print(f"  [yellow][Reporter][/yellow] [DEBUG] Project IDs: {project_ids[:5]}...")
        
        cost_data = []      # For Excel (Aggregated by Tech)
        project_data = []   # For Plotting (Granular Process + Tech)
        
        for t in self.years:
            row_agg = {'Year': t}
            row_proj = {'Year': t}
            
            # Aggregated row (Excel)
            for t_id in self.data.technologies:
                row_agg[t_id] = sum(tech_capex_spent.get((t, p_id, t_id), 0.0) for p_id in self.opt.entity.processes)
                # Group labels by tech
                all_labels = []
                for p_id in self.opt.entity.processes:
                    all_labels.extend(tech_invested_procs.get((t, p_id, t_id), []))
                row_agg[f"{t_id}_labels"] = ", ".join(sorted(list(set(all_labels))))
                row_agg[f"Aid_GRANT_{t_id}"] = aids_dist.get((t, f"Aid_GRANT_{t_id}"), 0.0)
                row_agg[f"Aid_CCFD_{t_id}"] = aids_dist.get((t, f"Aid_CCFD_{t_id}"), 0.0)
            
            if self.data.dac_params.active:
                row_agg['DAC'] = tech_capex_spent.get((t, 'INDIRECT', 'DAC'), 0.0)
                row_agg['DAC_labels'] = ", ".join(tech_invested_procs.get((t, 'INDIRECT', 'DAC'), []))
                dac_total_v = self._get_safe_var_value(self.opt.dac_total_capacity_vars.get(t))
                dac_captured_v = self._get_safe_var_value(self.opt.dac_captured_vars.get(t))
                base_opex = dac_total_v * self.data.dac_params.opex_by_year.get(t, 0.0)
                elec_cons = dac_captured_v * self.data.dac_params.elec_by_year.get(t, 0.0)
                elec_price = self.data.time_series.resource_prices.get('EN_ELEC', {}).get(t, 0.0)
                row_agg['DAC_Opex'] = base_opex + (elec_cons * elec_price)
            if self.data.credit_params.active:
                cred_v = self._get_safe_var_value(self.opt.credit_purchased_vars.get(t))
                row_agg['Credit_Cost'] = cred_v * self.data.credit_params.cost_by_year.get(t, 0.0)
            
            # Granular row (Plotting)
            for p_id, t_id in project_ids:
                col_name = f"{p_id}##{t_id}"
                row_proj[col_name] = tech_capex_spent.get((t, p_id, t_id), 0.0)
                row_proj[f"{col_name}##tCO2"] = tech_co2_abatement.get((t, p_id, t_id), 0.0)
                row_proj[f"{col_name}_is_new"] = 1 if (t, p_id, t_id) in new_investments else 0
                row_proj[f"{col_name}_labels"] = ", ".join(tech_invested_procs.get((t, p_id, t_id), []))

            cost_data.append(row_agg)
            project_data.append(row_proj)

        self.df_costs = pd.DataFrame(cost_data)
        self.df_projects = pd.DataFrame(project_data)
        df_costs = self.df_costs
        for t in self.years:
            # Compute Annual Budget Limit directly: CA revenue × ca_percentage_limit
            # This is independent of whether the solver created CA/Budget variables
            budget_val = 0.0
            ca_pct = self.opt.entity.ca_percentage_limit
            if t == self.years[0] and self.verbose:
                print(f"  [yellow][Reporter][/yellow] [DEBUG] Entity '{self.opt.entity.id}' budget limit factor: {ca_pct*100:.4f}%")
            if ca_pct > 0:
                # Try solver variable first
                ca_var_name = f"CA_{t}"
                yearly_var_name = f"YearlyBudget_{t}"
                vars_dict = self.opt.model.variablesDict()
                if ca_var_name in vars_dict and self._get_safe_var_value(vars_dict[ca_var_name], None) is not None:
                    budget_val = abs(self._get_safe_var_value(vars_dict[ca_var_name])) * ca_pct
                elif yearly_var_name in vars_dict and self._get_safe_var_value(vars_dict[yearly_var_name], None) is not None:
                    budget_val = self._get_safe_var_value(vars_dict[yearly_var_name])
                else:
                    # Fallback
                    for res_id in self.opt.entity.sold_resources:
                        price = self.data.time_series.resource_prices.get(res_id, {}).get(t, 0.0)
                        base_cons = abs(self.opt.entity.base_consumptions.get(res_id, 0.0))
                        budget_val += price * base_cons * ca_pct
            
            self.df_costs.loc[self.df_costs['Year'] == t, 'Budget_Limit'] = budget_val
            # Add Total Investment Limit (Budget + Loans, typically 1.5x Budget per optimizer logic)
            self.df_costs.loc[self.df_costs['Year'] == t, 'Total_Limit'] = 1.5 * budget_val
        
        # 1.6 Financing Data (Detailed Breakdown)
        financing_data = []
        annual_principal = {t: 0.0 for t in self.years}
        annual_interest = {t: 0.0 for t in self.years}
        # Track principal repayment per granular project to distribute back in the chart
        # Include 'INDIRECT' for DAC/Credit projects
        tech_principal_repayment = {(t, p_id, t_id): 0.0 for t in self.years for p_id in list(self.opt.entity.processes.keys()) + ['INDIRECT'] for t_id in list(self.data.technologies.keys()) + ['DAC', 'CREDIT']}
        
        loan_taken_per_year = {t: sum(self._get_safe_var_value(self.opt.loan_vars.get((t, p_id, t_id, l_id))) 
                                      for p_id, p in self.opt.entity.processes.items() 
                                      for t_id in p.valid_technologies if t_id != 'UP'
                                      for l_id in range(len(self.data.bank_loans))) for t in self.years}

        # First pass to calculate annual totals and technology shares for distribution
        for tau in self.years:
            # Calculate project shares for this year's investment
            project_cols = [f"{p_id}##{t_id}" for p_id, t_id in project_ids]
            total_capex_tau = self.df_projects.loc[self.df_projects['Year'] == tau, project_cols].sum(axis=1).values[0]
            project_shares = {}
            if total_capex_tau > 1e-4:
                for p_id, t_id in project_ids:
                    col_name = f"{p_id}##{t_id}"
                    project_shares[(p_id, t_id)] = self.df_projects.loc[self.df_projects['Year'] == tau, col_name].values[0] / total_capex_tau

            for l_id, loan in enumerate(self.data.bank_loans):
                p_initial = sum(self._get_safe_var_value(self.opt.loan_vars.get((tau, p_id, t_id, l_id)))
                                for p_id, p in self.opt.entity.processes.items()
                                for t_id in p.valid_technologies if t_id != 'UP')
                if p_initial > 0:
                    r = loan.rate
                    d = loan.duration
                    if r > 0:
                        annuity = p_initial * (r / (1 - (1 + r)**(-d)))
                    else:
                        annuity = p_initial / d
                    
                    current_balance = p_initial
                    for k in range(d):
                        t_rep = tau + k
                        if t_rep in self.years:
                            interest = current_balance * r
                            principal = annuity - interest
                            annual_principal[t_rep] += principal
                            annual_interest[t_rep] += interest
                            
                            # Distribute principal back to projects based on shares at time tau
                            for (p_id, t_id), share in project_shares.items():
                                tech_principal_repayment[(t_rep, p_id, t_id)] += principal * share
                                
                            current_balance -= principal

        # Construct financing dataframe and adjust cost data
        for t in self.years:
            tech_cols = [t_id for t_id in self.data.technologies if t_id != 'UP' and t_id in df_costs.columns]
            if self.data.dac_params.active and 'DAC' in df_costs.columns:
                tech_cols.append('DAC')
            total_capex = df_costs.loc[df_costs['Year'] == t, tech_cols].sum(axis=1).values[0]
            loan_taken = loan_taken_per_year[t]
            
            financing_data.append({
                'Year': t,
                'Total_CAPEX (M€)': total_capex / 1_000_000.0,
                'Loan_Principal_Taken (M€)': loan_taken / 1_000_000.0,
                'Out_of_Pocket_CAPEX (M€)': (total_capex - loan_taken) / 1_000_000.0,
                'Principal_Repayment (M€)': annual_principal[t] / 1_000_000.0,
                'Interest_Paid (M€)': annual_interest[t] / 1_000_000.0,
                'Total_Annuity (M€)': (annual_principal[t] + annual_interest[t]) / 1_000_000.0
            })
            
            # Adjust technology columns: Out-of-Pocket (current year) + Principal Repayment (from previous loans)
            for t_id in tech_cols:
                # 1. Current year out-of-pocket portion
                    if total_capex > 1e-4:
                        ratio_oop = (total_capex - loan_taken) / total_capex
                        current_oop = df_costs.loc[df_costs['Year'] == t, t_id].values[0] * ratio_oop
                    else:
                        current_oop = 0.0
                        
                    # 2. Add repayments for this technology due this year
                    repayment_portion = sum(tech_principal_repayment.get((t, p_id, t_id), 0.0) for p_id in self.opt.entity.processes)
                    
                    df_costs.loc[df_costs['Year'] == t, t_id] = current_oop + repayment_portion

            # The user requested that the Investment Plan chart only shows raw investment,
            # ignoring loan smoothing. So we do NOT modify self.df_projects here anymore.
            # self.df_projects retains the original `tech_capex_spent` raw values.
            # Add interest to df_costs for the stacked bar chart
            df_costs.loc[df_costs['Year'] == t, 'Financing Interests'] = annual_interest[t]
            # Also add to df_projects for consistency with the plotter's expectations
            self.df_projects.loc[self.df_projects['Year'] == t, 'Financing Interests'] = annual_interest[t]

        df_finance = pd.DataFrame(financing_data)
        
        # Aggregate CCfD Revenue per year for the Carbon Tax Graph
        ccfd_annual_revenue = {}
        for t in self.years:
            ccfd_annual_revenue[t] = sum(aids_dist.get((t, f"Aid_CCFD_{t_id}"), 0.0) for t_id in self.data.technologies) / 1_000_000.0
            
        df_emis['CCfD_Refund_MEuros'] = df_emis['Year'].map(ccfd_annual_revenue)
        df_emis['Net_Tax_Cost_MEuros'] = df_emis['Tax_Cost_MEuros'] - df_emis['CCfD_Refund_MEuros']
        
        # 1.6 Check for Missed Objectives (Soft Constraints)
        if self.verbose:
            print("\n  [yellow][Reporter][/yellow] [bold underline]--- Goal Achievement Report ---[/bold underline]")
        missed_goals = False
        for i, obj in enumerate(self.data.objectives):
            if obj.mode == 'LINEAR':
                t_target = min(obj.target_year, self.opt.years[-1])
                penalty_val = self._get_safe_var_value(self.opt.penalty_vars.get((i, t_target)))
                if penalty_val > 1e-4:
                    if obj.penalty_type == "NONE":
                        if self.verbose:
                            print(f"  [cyan][i] INFO: Objective {i+1} ('{obj.name}') not met at year {t_target} (Informational only: Gap {penalty_val:,.2f})[/i]")
                    else:
                        missed_goals = True
                        if self.verbose:
                            print(f"  [bold red][!] WARNING: Objective {i+1} ('{obj.name}') MISSED at year {t_target}![/bold red]")
                            print(f"    Target End: [cyan]{obj.target_year}[/cyan] | Resource: [cyan]{obj.resource}[/cyan] | Limit: [cyan]{obj.cap_value}[/cyan]")
                            print(f"    Shortfall (Penalty Paid): [bold red]{penalty_val:,.2f}[/bold red] {self.data.resources.get(obj.resource).unit if obj.resource in self.data.resources else 'units'}")
            else:
                penalty_val = self._get_safe_var_value(self.opt.penalty_vars.get(i))
                if penalty_val > 1e-4:
                    if obj.penalty_type == "NONE":
                        if self.verbose:
                            print(f"  [cyan][i] INFO: Objective {i+1} ('{obj.name}') not met (Informational only: Gap {penalty_val:,.2f})[/i]")
                    else:
                        missed_goals = True
                        if self.verbose:
                            print(f"  [bold red][!] WARNING: Objective {i+1} ('{obj.name}') MISSED![/bold red]")
                            print(f"    Target: [cyan]{obj.target_year}[/cyan] | Resource: [cyan]{obj.resource}[/cyan] | Limit: [cyan]{obj.cap_value}[/cyan]")
                            print(f"    Shortfall (Penalty Paid): [bold red]{penalty_val:,.2f}[/bold red] {self.data.resources.get(obj.resource).unit if obj.resource in self.data.resources else 'units'}")
        
        if not missed_goals:
            if self.verbose:
                print("  [bold green][OK] SUCCESS: All stated objectives were successfully met by the model.[/bold green]")
        if self.verbose:
            print("  [yellow][Reporter][/yellow] -------------------------------\n")
            
            # --- NEW: H2 Sourcing Breakdown Summary in Console ---
            h2_buy_cols = [c for c in df_cons.columns if 'H2' in c.upper() and c not in ['Year', 'EN_H2_C', 'EN_H2_P', 'EN_ELEC_FOR_H2']]
            if h2_buy_cols and (df_cons[h2_buy_cols] > 1.0).any().any():
                print("\n  >>> HYDROGEN SOURCING BREAKDOWN (Optimal Mix) <<<")
                for t in self.years:
                    row = df_cons[df_cons['Year'] == t]
                    total_h2 = row[h2_buy_cols].sum(axis=1).values[0]
                    if total_h2 > 1.0:
                        shares = []
                        for col in h2_buy_cols:
                            val = row[col].values[0]
                            if val > 1.0:
                                display_name = col.replace('EN_', '').replace('_C', '').replace('_', ' ')
                                if display_name == 'H2 ON SITE': display_name = 'ON-SITE ELECTROLYSIS'
                                shares.append(f"[cyan]{display_name}[/cyan]: {val:,.0f} units ({val/total_h2*100:.1f}%)")
                        if shares:
                            print(f"  Year {t}: {', '.join(shares)}")
                    print("")
        
        
        # 1.5 Prepare Data Used sheet
        data_used_rows = []
        # Get all unique resource IDs
        all_resources = set(self.data.time_series.resource_prices.keys()).union(set(self.data.time_series.other_emissions_factors.keys()))
        for r_id in all_resources:
            for t in self.years:
                price = self.data.time_series.resource_prices.get(r_id, {}).get(t, np.nan)
                emissions = self.data.time_series.other_emissions_factors.get(r_id, {}).get(t, np.nan)
                
                res = self.data.resources.get(r_id)
                if res and res.unit.upper() == 'GJ':
                    # If consumption was divided by 3.6, price and emissions must be multiplied by 3.6
                    if not np.isnan(price): price *= 3.6
                    if not np.isnan(emissions): emissions *= 3.6
                
                data_used_rows.append({
                    "Resource": r_id,
                    "Year": t,
                    "Price": price,
                    "CO2_Emissions": emissions
                })
        df_data_used = pd.DataFrame(data_used_rows) if data_used_rows else pd.DataFrame(columns=["Resource", "Year", "Price", "CO2_Emissions"])

        # 2. Export to Excel
        excel_path = os.path.join(self.results_dir, 'Master_Plan.xlsx')
        if self.generate_excel or (hasattr(self.data, 'reporting_toggles') and self.data.reporting_toggles.results_excel):
            try:
                with pd.ExcelWriter(excel_path) as writer:
                    if not df_invest.empty:
                        df_invest.to_excel(writer, sheet_name='Investments', index=False)
                    df_cons.to_excel(writer, sheet_name='Energy_Mix', index=False)
                    df_emis.to_excel(writer, sheet_name='CO2_Trajectory', index=False)
                    if not df_indir.empty:
                        df_indir.to_excel(writer, sheet_name='Indirect_Emissions', index=False)
                    df_costs.to_excel(writer, sheet_name='Technology_Costs', index=False)
                    df_finance.to_excel(writer, sheet_name='Financing', index=False)
                    
                    if not df_data_used.empty:
                        df_data_used.to_excel(writer, sheet_name='Data_Used', index=False)
                    
                    # NEW: Export Resource Metadata for Dashboard Categories
                    res_metadata = []
                    for r_id, res in self.data.resources.items():
                        unit = res.unit
                        if unit.upper() == 'GJ':
                            unit = 'MWh'
                        res_metadata.append({
                            "ID": r_id,
                            "Name": res.name,
                            "Type": res.type,
                            "Category": res.category,
                            "Unit": unit
                        })
                    if res_metadata:
                        pd.DataFrame(res_metadata).to_excel(writer, sheet_name='Resource_Metadata', index=False)
                    
                    # 2.2 Export Chart Data
                    if hasattr(self, 'charts_data') and self.charts_data:
                        start_row = 0
                        sheet_name = 'Charts'
                        for title, df_chart in self.charts_data:
                            # Write Title
                            pd.Series([title]).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False)
                            # Write Data
                            df_chart.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 1, index=True)
                            # Advance 2 blank lines (title + header + data rows + 2)
                            start_row += len(df_chart) + 4
                    
                if self.verbose:
                    print(f"  [yellow][Reporter][/yellow] [OK] Exported data to [bold cyan]{excel_path}[/bold cyan]")
            except PermissionError:
                os.makedirs(results_dir, exist_ok=True)
                excel_path = os.path.join(results_dir, 'Master_Plan_new.xlsx')
                with pd.ExcelWriter(excel_path) as writer:
                    if not df_invest.empty:
                        df_invest.to_excel(writer, sheet_name='Investments', index=False)
                    df_cons.to_excel(writer, sheet_name='Energy_Mix', index=False)
                    df_emis.to_excel(writer, sheet_name='CO2_Trajectory', index=False)
                    if not df_indir.empty:
                        df_indir.to_excel(writer, sheet_name='Indirect_Emissions', index=False)
                    df_costs.to_excel(writer, sheet_name='Technology_Costs', index=False)
                    df_finance.to_excel(writer, sheet_name='Financing', index=False)
    
                    if not df_data_used.empty:
                        df_data_used.to_excel(writer, sheet_name='Data_Used', index=False)
                    
                    # NEW: Export Resource Metadata for Dashboard Categories (Fallback)
                    res_metadata = []
                    for r_id, res in self.data.resources.items():
                        unit = res.unit
                        if unit.upper() == 'GJ':
                            unit = 'MWh'
                        res_metadata.append({
                            "ID": r_id,
                            "Name": res.name,
                            "Type": res.type,
                            "Category": res.category,
                            "Unit": unit
                        })
                    if res_metadata:
                        pd.DataFrame(res_metadata).to_excel(writer, sheet_name='Resource_Metadata', index=False)
                        
                    # 2.2 Export Chart Data (Fallback)
                    if hasattr(self, 'charts_data') and self.charts_data:
                        start_row = 0
                        sheet_name = 'Charts'
                        for title, df_chart in self.charts_data:
                            pd.Series([title]).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False)
                            df_chart.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 1, index=True)
                            start_row += len(df_chart) + 4
    
                if self.verbose:
                    print(f"  [yellow][Reporter][/yellow] [!] File was locked. Exported data to [bold cyan]{excel_path}[/bold cyan] instead.")
        # 3. NEW: CO2 Abatement Cost Calculation (MAC) - Aggregate by Tech Type
        # ── 1. HELPERS: Uniform effective price and emission factor ──────────────────────
        # ALL resources get the same treatment:
        #   effective_price = market_price + emission_factor × carbon_tax  (Scope 3 included)
        #   effective_emiss = emission_factor from time series
        # No hardcoded resource IDs — the logic reads whatever is in the data.

        def _get_effective_price(r_id, yr):
            """All-in cost for one unit of resource r_id at year yr.
            Includes the market price plus the indirect carbon cost of upstream emissions."""
            market_p = self.data.time_series.resource_prices.get(r_id, {}).get(yr, 0.0)
            emiss_f  = self.data.time_series.other_emissions_factors.get(r_id, {}).get(yr, 0.0)
            carbon_p = self.data.time_series.carbon_prices.get(yr, 0.0)
            return market_p + (emiss_f * carbon_p)

        def _get_effective_emiss(r_id, yr):
            """Indirect emission factor (tCO2/unit) for resource r_id at year yr."""
            return self.data.time_series.other_emissions_factors.get(r_id, {}).get(yr, 0.0)

        def _get_metric(r_id, yr, metric_type):
            """Unified metric getter — same logic for every resource."""
            if metric_type == 'price':
                return _get_effective_price(r_id, yr)
            else:
                return _get_effective_emiss(r_id, yr)

        def _primary_energy_consumption(process):
            """Return the base consumption of the dominant energy resource for a process.
            Searches all base_consumptions ranked by share — avoids hardcoding 'EN_FUEL'."""
            best_res, best_val = None, 0.0
            for res, share in process.consumption_shares.items():
                base = self.opt.entity.base_consumptions.get(res, 0.0)
                val  = base * share
                if val > best_val:
                    best_val = val
                    best_res = res
            return best_val, best_res

        # ── 2. AGGREGATE MAC BY TECHNOLOGY ───────────────────────────────────────────────
        tech_mac_agg = {}   # t_id → {capex, opex, co2, processes, status}

        all_techs = [t_id for t_id in self.data.technologies.keys() if t_id != 'UP']
        if self.data.dac_params.active:    all_techs.append('DAC')
        if self.data.credit_params.active: all_techs.append('CREDIT')

        for t_id in all_techs:
            if t_id not in tech_mac_agg:
                tech_mac_agg[t_id] = {'capex': 0.0, 'opex': 0.0, 'co2': 0.0,
                                      'processes': [], 'status': 'Potential'}

            if   t_id == 'DAC':    valid_p_ids = ['INDIRECT']
            elif t_id == 'CREDIT': valid_p_ids = ['INDIRECT']
            else:
                valid_p_ids = [p_id for p_id, p in self.opt.entity.processes.items()
                               if t_id in p.valid_technologies]

            is_invested = any((p_id, t_id) in project_ids for p_id in valid_p_ids)

            if is_invested:
                # ── INVESTED path ────────────────────────────────────────────
                tech_mac_agg[t_id]['status'] = 'Invested'
                tech = self.data.technologies.get(t_id)

                for p_id in valid_p_ids:
                    if (p_id, t_id) not in project_ids:
                        continue
                    process = self.opt.entity.processes.get(p_id)
                    if not tech or not process:
                        continue

                    # A — CAPEX (actual spent)
                    tech_mac_agg[t_id]['capex'] += sum(
                        tech_capex_spent.get((t, p_id, t_id), 0.0) for t in self.years)

                    # B — CO2 Abated (uses actual activation variables)
                    for res_id, imp in tech.impacts.items():
                        for t in self.years:
                            act_var = getattr(self.opt.active_vars.get((t, p_id, t_id)),
                                              'varValue', 0.0) or 0.0
                            if act_var < 1e-6:
                                continue

                            if res_id == 'CO2_EM' or 'CO2' in res_id.upper():
                                y_factor = 1.0
                            else:
                                y_factor = _get_effective_emiss(res_id, t)
                                if y_factor == 0.0:
                                    continue

                            if self._is_primary_emission_resource(res_id):
                                base_val = self.opt.entity.process_emission_baseline(process, primary_emis)
                            else:
                                base_val = (self.opt.entity.base_consumptions.get(res_id, 0.0)
                                            * process.consumption_shares.get(res_id, 0.0))

                            yr_idx  = list(self.opt.years).index(t)
                            up_rate = 0.0
                            if 'UP' in process.valid_technologies:
                                up_tech = self.data.technologies['UP']
                                up_imp  = up_tech.impacts.get('ALL') or up_tech.impacts.get(res_id)
                                if up_imp and up_imp['type'] in ['variation', 'up']:
                                    up_rate = abs(up_imp['value'])
                            base_val *= (1.0 - up_rate) ** yr_idx

                            if imp['type'] in ['variation', 'up']:
                                target_red = -imp['value'] * base_val
                            elif imp['type'] == 'new':
                                ref_res  = imp.get('ref_resource')
                                if self._is_primary_emission_resource(ref_res):
                                    base_ref = self.opt.entity.process_emission_baseline(process, ref_res)
                                elif ref_res and ref_res in self.opt.entity.base_consumptions:
                                    base_ref = (self.opt.entity.base_consumptions.get(ref_res, 0.0)
                                                * process.consumption_shares.get(ref_res, 0.0))
                                else:
                                    base_ref = base_val
                                target_red = -imp['value'] * base_ref
                            else:
                                target_red = 0.0

                            tech_mac_agg[t_id]['co2'] += (target_red * y_factor) * (act_var / process.nb_units)

                    # C — OPEX Change (actual, summed over all active years)
                    for t in self.years:
                        act_var = getattr(self.opt.active_vars.get((t, p_id, t_id)),
                                          'varValue', 0.0) or 0.0
                        if act_var < 1e-6:
                            continue

                        year_opex_change = 0.0
                        for res_id, imp_op in tech.impacts.items():
                            if res_id in ('CO2_EM', 'ALL'):
                                continue

                            ref_res   = imp_op.get('ref_resource') or res_id
                            ref_state = imp_op.get('reference', 'INITIAL')

                            if self._is_primary_emission_resource(ref_res):
                                base_ref = self.opt.entity.process_emission_baseline(process, primary_emis)
                            else:
                                base_ref = (self.opt.entity.base_consumptions.get(ref_res, 0.0)
                                            * process.consumption_shares.get(ref_res, 0.0))

                            if imp_op['type'] in ('variation', 'up'):
                                target_change = imp_op['value'] * base_ref
                            elif imp_op['type'] == 'new':
                                if ref_state == 'AVOIDED':
                                    co2_imp = tech.impacts.get(primary_emis)
                                    if not co2_imp:
                                        imp_keys = [k for k in tech.impacts.keys() if self._is_primary_emission_resource(k)]
                                        if imp_keys: co2_imp = tech.impacts[imp_keys[0]]

                                    if co2_imp and co2_imp['type'] == 'variation' and (self._is_primary_emission_resource(ref_res) or not ref_res):
                                        target_change = imp_op['value'] * abs(co2_imp['value'] * base_ref)
                                    else:
                                        target_change = imp_op['value'] * base_ref
                                else:
                                    target_change = imp_op['value'] * base_ref
                            else:
                                target_change = 0.0

                            year_opex_change += (target_change * (act_var / process.nb_units)) \
                                                * _get_effective_price(res_id, t)

                        current_opex = tech.opex_by_year.get(t, tech.opex)
                        cap_opex     = 1.0
                        if tech.opex_per_unit:
                            if tech.opex_unit == 'tCO2':
                                cap_opex = self.opt.entity.process_emission_baseline(process, self._primary_emission_resource())
                            else:
                                primary_conso, _ = process.primary_energy_consumption(self.opt.entity.base_consumptions)
                                cap_opex = primary_conso / self.opt.entity.annual_operating_hours

                        tech_mac_agg[t_id]['opex'] += (
                            year_opex_change + (current_opex * cap_opex / process.nb_units) * act_var)

                    tech_mac_agg[t_id]['processes'].append(p_id)

            else:
                # ── POTENTIAL path ───────────────────────────────────────────
                t_eval = self.years[0]
                tech   = self.data.technologies.get(t_id)
                if not tech:
                    continue

                for p_id in valid_p_ids:
                    process = self.opt.entity.processes.get(p_id)
                    if not process:
                        continue

                    # A — Potential CAPEX
                    cap_capex = 1.0
                    if tech.capex_per_unit:
                        if tech.capex_unit == 'tCO2':
                            cap_capex = self.opt.entity.process_emission_baseline(process, self._primary_emission_resource())
                        elif 'MW' in tech.capex_unit.upper():
                            primary_conso, _ = _primary_energy_consumption(process)
                            cap_capex = primary_conso / self.opt.entity.annual_operating_hours
                    tech_mac_agg[t_id]['capex'] += tech.capex_by_year.get(t_eval, tech.capex) * cap_capex

                    # B — Potential CO2 Abated
                    for res_id, imp in tech.impacts.items():
                        if res_id == 'CO2_EM' or 'CO2' in res_id.upper():
                            factor = 1.0
                        else:
                            factor = _get_effective_emiss(res_id, t_eval)
                            if factor == 0.0:
                                continue

                        if self._is_primary_emission_resource(res_id):
                            base_val = self.opt.entity.process_emission_baseline(process, primary_emis)
                        else:
                            base_val = (self.opt.entity.base_consumptions.get(res_id, 0.0)
                                        * process.consumption_shares.get(res_id, 0.0))

                        if imp['type'] in ('variation', 'up'):
                            target_red = -imp['value'] * base_val
                        elif imp['type'] == 'new':
                            ref_res  = imp.get('ref_resource')
                            if self._is_primary_emission_resource(ref_res):
                                base_ref = self.opt.entity.process_emission_baseline(process, primary_emis)
                            elif ref_res and ref_res in self.opt.entity.base_consumptions:
                                base_ref = (self.opt.entity.base_consumptions.get(ref_res, 0.0)
                                            * process.consumption_shares.get(ref_res, 0.0))
                            else:
                                base_ref = base_val
                            target_red = -imp['value'] * base_ref
                        else:
                            target_red = 0.0

                        tech_mac_agg[t_id]['co2'] += (target_red * factor) \
                                                     * (len(self.years) - tech.implementation_time)

                    # C — Potential OPEX Change
                    for t in self.years[tech.implementation_time:]:
                        year_opex_change = 0.0
                        for res_id, imp_op in tech.impacts.items():
                            if res_id in ('CO2_EM', 'ALL'):
                                continue

                            ref_res   = imp_op.get('ref_resource') or res_id
                            ref_state = imp_op.get('reference', 'INITIAL')

                            if self._is_primary_emission_resource(ref_res):
                                base_ref = self.opt.entity.process_emission_baseline(process, primary_emis)
                            else:
                                base_ref = (self.opt.entity.base_consumptions.get(ref_res, 0.0)
                                            * process.consumption_shares.get(ref_res, 0.0))

                            if imp_op['type'] in ('variation', 'up'):
                                target_change = imp_op['value'] * base_ref
                            elif imp_op['type'] == 'new':
                                if ref_state == 'AVOIDED':
                                    co2_imp = tech.impacts.get(primary_emis)
                                    if not co2_imp:
                                        imp_keys = [k for k in tech.impacts.keys() if self._is_primary_emission_resource(k)]
                                        if imp_keys: co2_imp = tech.impacts[imp_keys[0]]
                                    
                                    if co2_imp and co2_imp['type'] == 'variation' and (self._is_primary_emission_resource(ref_res) or not ref_res):
                                        target_change = imp_op['value'] * abs(co2_imp['value'] * base_ref)
                                    else:
                                        target_change = imp_op['value'] * base_ref
                                else:
                                    target_change = imp_op['value'] * base_ref
                            else:
                                target_change = 0.0

                            year_opex_change += target_change * _get_effective_price(res_id, t)

                        cur_opex = tech.opex_by_year.get(t, tech.opex)
                        cap_opex = 1.0
                        if tech.opex_per_unit:
                            if tech.opex_unit == 'tCO2':
                                cap_opex = self.opt.entity.process_emission_baseline(process, primary_emis)
                            elif 'MW' in tech.opex_unit.upper():
                                primary_conso = self._process_primary_energy_consumption(process)
                                cap_opex = primary_conso / self.opt.entity.annual_operating_hours

                        tech_mac_agg[t_id]['opex'] += year_opex_change + (cur_opex * cap_opex)

                    tech_mac_agg[t_id]['processes'].append(p_id)

        # DAC and CREDIT special potential logic
        if 'DAC' in tech_mac_agg and tech_mac_agg['DAC']['status'] == 'Potential':
            tech_mac_agg['DAC']['co2']   = 100_000.0 * (len(self.years) - 2)
            tech_mac_agg['DAC']['capex'] = 100_000.0 * self.data.dac_params.capex_by_year.get(self.years[0], 500.0)
            tech_mac_agg['DAC']['opex']  = 100_000.0 * self.data.dac_params.opex_by_year.get(self.years[0], 100.0) * (len(self.years) - 2)
        if 'CREDIT' in tech_mac_agg and tech_mac_agg['CREDIT']['status'] == 'Potential':
            tech_mac_agg['CREDIT']['co2']  = 50_000.0 * len(self.years)
            tech_mac_agg['CREDIT']['opex'] = 50_000.0 * self.data.credit_params.cost_by_year.get(self.years[0], 50.0) * len(self.years)

        # ── 3. BUILD MAC DATA LIST ────────────────────────────────────────────────────────
        # One entry per technology ID — no decomposition by sub-resource.
        # Each FUEL_TO_H2_X variant is already its own t_id and produces its own bar.
        mac_data = []
        for t_id, vals in tech_mac_agg.items():
            if vals['co2'] < 1.0:
                continue

            if t_id in self.data.technologies:
                tech_name = self.data.technologies[t_id].name
            elif t_id == 'DAC':
                tech_name = "Direct Air Capture"
            elif t_id == 'CREDIT':
                tech_name = "Carbon Credits"
            else:
                tech_name = t_id

            proc_sum = ", ".join(sorted(set(vals['processes']))[:3])
            if len(set(vals['processes'])) > 3:
                proc_sum += ", ..."

            mac_data.append({
                'Project':               tech_name,
                'Process Summary':       proc_sum,
                'Display Label':         f"{tech_name}\n({proc_sum})" if proc_sum else tech_name,
                'MAC (€/tCO2)':          (vals['capex'] + vals['opex']) / vals['co2'],
                'MAC CAPEX (€/tCO2)':    vals['capex'] / vals['co2'],
                'MAC OPEX (€/tCO2)':     vals['opex']  / vals['co2'],
                'Total Abated (tCO2)':   vals['co2'],
                'Total CAPEX (M€)':      vals['capex'] / 1_000_000.0,
                'Total OPEX Change (M€)': vals['opex'] / 1_000_000.0,
                'Status':                vals['status'],
            })

        # Build MAC dataframe once so chart gating can safely test emptiness.
        df_mac = pd.DataFrame(mac_data)
        if not df_mac.empty:
            df_mac = df_mac.sort_values(by='MAC (€/tCO2)').reset_index(drop=True)

        # 4. Generate Visualizations (Always export JSON for Dashboard, gate PNG via show_png)
        if self.verbose:
            print("  [yellow][Reporter][/yellow] [PLOT] Generating visualizations...")
        self.charts_data = [] # Clear/Re-init for collection
        toggles = self.data.reporting_toggles

        # Always process all key visualizations to populate the Dashboard artifacts
        self._plot_resources_mix(df_cons, df_indir, show_png=toggles.chart_energy_mix)
        _step()
        
        self._plot_co2_trajectory(df_emis, show_png=toggles.chart_co2_trajectory)
        _step()
        
        self._plot_indirect_emissions(df_indir, show_png=toggles.chart_indirect_emissions)
        _step()
        
        self._plot_investment_costs(df_costs, df_finance, show_png=toggles.chart_investment_costs)
        _step()
        
        self._plot_total_opex(df_cons, df_emis, show_png=toggles.chart_total_opex)
        _step()
        
        self._plot_carbon_tax_and_avoided(df_emis, show_png=toggles.chart_carbon_tax_avoided)
        _step()
        
        self._plot_external_financing(df_costs, df_finance, show_png=toggles.chart_external_financing)
        _step()
        
        self._plot_transition_costs(df_costs, df_finance, df_emis, show_png=toggles.chart_transition_costs)
        _step()
        
        self._plot_carbon_prices(show_png=toggles.chart_carbon_prices)
        _step()
        
        self._plot_interest_paid(df_finance, show_png=toggles.chart_interest_paid)
        _step()
        
        self._plot_prices(show_png=toggles.chart_resource_prices)
        _step()
        
        self._plot_co2_abatement_cost(df_mac, show_png=toggles.chart_co2_abatement_cost)
        _step()
        
        # 4. Final step: Re-run Excel export to include the charts sheet if needed
        # (This is a bit redundant but ensures we have the data captured AFTER plots are run)
        if self.generate_excel:
            self._export_charts_sheet(excel_path)
            _step()

    def _export_charts_sheet(self, excel_path: str):
        """Append or overwrite the 'Charts' sheet with collected plot data."""
        if not self.charts_data:
            return
            
        try:
            # Check if file exists to decide mode
            if os.path.exists(excel_path):
                # To avoid pandas overwrite loop bugs with replace, remove the sheet first
                from openpyxl import load_workbook
                wb = load_workbook(excel_path)
                if 'Charts' in wb.sheetnames:
                    del wb['Charts']
                    wb.save(excel_path)
                wb.close()
                with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    start_row = 0
                    sheet_name = 'Charts'
                    for title, df_chart in self.charts_data:
                        if df_chart is None or df_chart.empty:
                            continue
                        pd.Series([title]).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False)
                        df_chart.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 1, index=True)
                        start_row += len(df_chart) + 4
            else:
                with pd.ExcelWriter(excel_path) as writer:
                    start_row = 0
                    sheet_name = 'Charts'
                    for title, df_chart in self.charts_data:
                        if df_chart is None or df_chart.empty:
                            continue
                        pd.Series([title]).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False)
                        df_chart.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 1, index=True)
                        start_row += len(df_chart) + 4
        except Exception as e:
            print(f"  [yellow][Reporter][/yellow] [!] Could not export charts data: {e}")

    def _plot_resources_mix(self, df_cons: pd.DataFrame, df_indir: pd.DataFrame, show_png: bool = True):
        """Generates the Resources Mix visualization with dynamic Type-Category grouping."""
        years = list(self.years)
        # We will now use a structure: { (TYPE, CATEGORY): { 'unit': unit, 'series': { name: values } } }
        # This allows dynamic creation of buttons in the plot builder.
        group_data = {}
        GJ_TO_MWH = 1.0 / 3.6

        def add_to_group(t_id, cat_name, name, unit, values, original_res_id=None):
            key = (t_id, cat_name)
            
            # 1. Determine common unit for this group (favor MWh for Energy categories)
            if key not in group_data:
                target_unit = unit
                # Heuristic: force MWh for any category containing 'ENERGY' if MWh is available
                if 'ENERGY' in str(cat_name).upper() and unit.upper() == 'GJ':
                    target_unit = 'MWh'
                group_data[key] = {'unit': target_unit, 'series': {}}
            
            target_unit = group_data[key]['unit']
            
            # 2. Normalize values if unit mismatch
            if unit.upper() != target_unit.upper():
                factor = self.data.unit_conversions.get((unit.upper(), target_unit.upper()), 1.0)
                # Fallback for GJ if not in conversions (1 MWh = 3.6 GJ)
                if factor == 1.0:
                    if unit.upper() == 'GJ' and target_unit.upper() == 'MWH': factor = 1.0 / 3.6
                    elif unit.upper() == 'MWH' and target_unit.upper() == 'GJ': factor = 3.6
                
                if factor != 1.0:
                    values = [v * factor for v in values]
            
            group_data[key]['series'][name] = values

        # 1. Process Consumption & Production
        df_c = df_cons.set_index('Year') if 'Year' in df_cons.columns else df_cons
        for col in df_c.columns:
            if col == 'Year': continue

            if col not in self.data.resources: continue
            res = self.data.resources[col]
            
            # Categorize based on Resource Type
            res_type = str(res.type).upper()
            if 'EMISS' in res_type: 
                t_id = 'EMISSIONS'
            elif 'PROD' in res_type:
                t_id = 'PRODUCTION'
            else:
                t_id = 'CONSUMPTION'
            
            # Only include energy-related or requested resources for this specific mix
            if not (col.startswith('EN_') or t_id != 'CONSUMPTION'):
                continue

            name = res.name if res.name else col
            unit = res.unit if res.unit else 'unit'
            vals = df_c[col].tolist()
            
            add_to_group(t_id, res.category, name, unit, vals, original_res_id=col)

        # 2. Process Indirect Emissions
        df_i = df_indir.set_index('Year') if 'Year' in df_indir.columns else df_indir
        for col in df_i.columns:
            if col == 'Year' or df_i[col].sum() < 1e-6: continue
            
            # Use resource info if column is a resource ID
            res_id = col
            if '_' in col and col not in self.data.resources:
                res_id = col.split('_')[0] + '_' + col.split('_')[1]
            
            res = self.data.resources.get(res_id)
            name = col.replace('_', ' ')
            cat_name = res.category if res else 'Indirect'
            unit = 'tCO2'
            vals = df_i[col].tolist()
            
            add_to_group('EMISSIONS', cat_name, name, unit, vals)

        # 3. Build figure
        fig = build_resources_mix_figure(
            years=years,
            group_data=group_data,
            theme="report",
            title="RESOURCES MIX: TYPE & CATEGORY"
        )
        
        self._save_plotly_figure(fig, "Energy_Mix", show_png=show_png)
        self.charts_data.append(("Resources Mix Breakdown (Type & Category Support)", df_cons))

    def _plot_co2_trajectory(self, df: pd.DataFrame, show_png: bool = True):
        """Generates the CO2 Trajectory visualization using the centralized module."""
        fig = build_co2_trajectory_figure(
            df=df,
            objectives=self.data.objectives,
            base_emissions=self.opt.entity.base_emissions,
            theme="report",
            title=f"CO2 EMISSIONS TRAJECTORY: {self.scenario_name}"
        )
        
        self._save_plotly_figure(fig, "CO2_Trajectory", show_png=show_png)
        self.charts_data.append(("CO2 Emissions Trajectory & Decarbonization Goals", df))

    def _plot_indirect_emissions(self, df: pd.DataFrame, show_png: bool = True):
        """Plot indirect emissions breakdown using the centralized module."""
        df_plot = df.copy()
        if 'Year' in df_plot.columns:
            df_plot.set_index('Year', inplace=True)
        # Remove resources that are all zeros
        df_plot = df_plot.loc[:, (df_plot != 0).any(axis=0)]
        
        if df_plot.empty:
            self._save_plotly_placeholder(
                "Indirect_Emissions", 
                "INDIRECT EMISSIONS", 
                "No indirect emissions data available for this scenario.",
                show_png=show_png
            )
            return

        # Convert to ktCO2
        df_plot = df_plot / 1000.0

        # Group by dynamic resource categories from metadata.
        cat_df = pd.DataFrame(index=df_plot.index)
        for col in df_plot.columns:
            res_id = col
            if '_' in col and col not in self.data.resources:
                parts = col.split('_')
                if len(parts) >= 2:
                    res_id = f"{parts[0]}_{parts[1]}"

            res = self.data.resources.get(res_id)
            group_name = res.name if res and res.name else (res.id if res else col.replace('_', ' '))

            if group_name not in cat_df.columns:
                cat_df[group_name] = 0.0
            cat_df[group_name] = cat_df[group_name] + df_plot[col]
            
        fig = build_indirect_emissions_figure(
            df_cat=cat_df,
            years=list(cat_df.index),
            theme="report",
            title=f"INDIRECT EMISSIONS: {self.scenario_name}"
        )
        
        self._save_plotly_figure(fig, "Indirect_Emissions", show_png=show_png)
        self.charts_data.append(("Indirect Emissions Breakdown by Resource (Scope 2 & 3)", cat_df))
        
    def _plot_investment_costs(self, df: pd.DataFrame, df_finance: pd.DataFrame = None, show_png: bool = True):
        """Plot the Investment Plan, focusing on Implementation Costs (M€) and budget limits using Plotly."""
        # 1. Prepare Data using the high-fidelity builder
        fig = build_investment_plan_figure(
            df_projects=self.df_projects,
            df_costs=self.df_costs,
            years=list(self.years),
            theme="report",
            title=f"INVESTMENT PLAN: {self.scenario_name}"
        )

        # 2. Save Results (PNG + JSON)
        self._save_plotly_figure(fig, "Investment_Plan", show_png=show_png)

        # 3. Store data for Excel
        excluded_suffixes = ('##tCO2', '_labels', '_is_new', 'Financing Interests', 'Year', 'Yearly_Total')
        capex_cols = [c for c in self.df_projects.columns if not any(c.endswith(s) for s in excluded_suffixes) and c != 'Year']
        df_plot = self.df_projects[['Year'] + capex_cols].copy()
        for col in self.df_projects.columns:
            if col.endswith('_labels'):
                df_plot[col] = self.df_projects[col]

        self.charts_data.append(("INVESTMENT_PLAN_HIGH_FIDELITY", df_plot))
        
        # Defensive check before set_index
        if 'Year' in df_plot.columns:
            df_summary = df_plot.set_index('Year')[capex_cols] / 1_000_000.0
            self.charts_data.append(("Investment Plan: Implementation Costs", df_summary))

    def _plot_external_financing(self, df_costs: pd.DataFrame, df_finance: pd.DataFrame, show_png: bool = True):
        """Standardizes the Public Aids chart using the high-fidelity builder."""
        df_costs = df_costs.copy()
        if 'Year' in df_costs.columns:
            df_costs.set_index('Year', inplace=True)
            
        # 1. Public Aids
        aid_cols = [c for c in df_costs.columns if c.startswith('Aid_')]
        df_aids = df_costs[aid_cols].copy()
        df_aids = df_aids.loc[:, (df_aids != 0).any(axis=0)]
        df_aids = df_aids / 1_000_000.0 # Convert to M€
        
        # 2. Private Financing (Bank Loans)
        df_fin = df_finance.copy()
        if 'Year' in df_fin.columns:
            df_fin.set_index('Year', inplace=True)
        private_loans = df_fin['Loan_Principal_Taken (M€)'] if 'Loan_Principal_Taken (M€)' in df_fin.columns else pd.Series(0.0, index=df_fin.index)
        
        # Merge for plotting
        df_plot_data = df_aids.copy()
        if private_loans.any():
            df_plot_data['Private Bank Loans'] = private_loans

        if df_plot_data.empty:
            self._save_plotly_placeholder(
                "Financing",
                "FINANCING STRATEGY",
                "No external financing or public aids data available for this scenario.",
                show_png=show_png
            )
            return

        # Rename columns for clarity
        rename_map = {}
        for col in df_plot_data.columns:
            if col.startswith('Aid_'):
                parts = col.split('_', 2)
                if len(parts) == 3:
                    t_name = self.data.technologies[parts[2]].name if parts[2] in self.data.technologies else parts[2]
                    rename_map[col] = f"{parts[1]} - {t_name}"
            else:
                rename_map[col] = col
        df_plot_data = df_plot_data.rename(columns=rename_map)

        fig = build_external_financing_figure(
            df_plot=df_plot_data,
            years=list(df_plot_data.index),
            theme="report",
            title=f"FINANCING STRATEGY: {self.scenario_name}"
        )
        
        self._save_plotly_figure(fig, "Financing", show_png=show_png)
        self.charts_data.append(("Generalized Financing Breakdown", df_plot_data))

    def _plot_total_opex(self, df_cons: pd.DataFrame, df_emis: pd.DataFrame, show_png: bool = True):
        """Unified Plotly visualization for OPEX breakdown using the centralized module."""
        years = list(self.opt.years)
        
        # 1. Prepare Data (Internal Calculation parity)
        df_cons_plot = df_cons.copy()
        if 'Year' in df_cons_plot.columns:
            df_cons_plot.set_index('Year', inplace=True)
        df_res_costs = df_cons_plot / 1_000_000.0
        tech_opex_details = {}
        for t_id, tech in self.data.technologies.items():
            if tech.is_continuous_improvement:
                continue
            tech_annual_opex = []
            is_active = False
            for t in years:
                yr_val = 0.0
                for p_id, process in self.opt.entity.processes.items():
                    if t_id in process.valid_technologies:
                        act_v = getattr(self.opt.active_vars[(t, p_id, t_id)], 'varValue', 0.0) or 0.0
                        if act_v > 0.01:
                            is_active = True
                            cap_opex = 1.0
                            if tech.opex_per_unit:
                                if tech.opex_unit == 'tCO2': 
                                    cap_opex = self.opt.entity.process_emission_baseline(process, self._primary_emission_resource())
                                elif 'MW' in str(tech.opex_unit).upper():
                                    best_val = self._process_primary_energy_consumption(process)
                                    cap_opex = best_val / self.opt.entity.annual_operating_hours
                            current_opex = tech.opex_by_year.get(t, tech.opex)
                            yr_val += (current_opex * cap_opex / process.nb_units) * act_v
                tech_annual_opex.append(yr_val / 1_000_000.0)
            if is_active:
                col_name = f"{tech.name.upper()} OPEX" if tech.name else f"{t_id.upper()} OPEX"
                tech_opex_details[col_name] = tech_annual_opex

        df_tech_costs = pd.DataFrame(tech_opex_details, index=years)
        
        dac_opex = []
        if self.data.dac_params.active:
            for t in years:
                dac_total_v = self._get_safe_var_value(self.opt.dac_total_capacity_vars.get(t))
                dac_opex.append((dac_total_v * self.data.dac_params.opex_by_year.get(t, 0.0)) / 1_000_000.0)
        
        df_emis_plot = df_emis.copy()
        if 'Year' in df_emis_plot.columns:
            df_emis_plot.set_index('Year', inplace=True)
        
        tax_costs = df_emis_plot['Tax_Cost_MEuros'].tolist() if 'Tax_Cost_MEuros' in df_emis_plot.columns else [0.0]*len(years)
        credit_costs = []
        for t in years:
            cred_v = self._get_safe_var_value(self.opt.credit_purchased_vars.get(t))
            credit_costs.append((cred_v * self.data.credit_params.cost_by_year.get(t, 0.0)) / 1_000_000.0)
            
        ccs_st_costs = []
        for t in years:
            yr_ccs_st = 0.0
            for p_id, process in self.opt.entity.processes.items():
                for t_id in process.valid_technologies:
                    tech = self.data.technologies[t_id]
                    if tech.tech_category == 'Carbon Capture':
                        primary_emis = self._primary_emission_resource()
                        imp = tech.impacts.get(primary_emis)
                        if not imp:
                            co2_keys = [k for k in tech.impacts.keys() if self._is_primary_emission_resource(k)]
                            if co2_keys: imp = tech.impacts[co2_keys[0]]
                            
                        if imp and (imp['type'] == 'variation' or imp['type'] == 'up'):
                            act_v = getattr(self.opt.active_vars[(t, p_id, t_id)], 'varValue', 0.0) or 0.0
                            if act_v > 0.01:
                                reduction_frac_per_unit = (-imp['value'] if imp['value'] < 0 else 0) / process.nb_units
                                max_emis_for_proc = self.opt.entity.process_emission_baseline(process, primary_emis)
                                captured_tons_per_unit = reduction_frac_per_unit * max_emis_for_proc
                                
                                s_price = self.data.time_series.resource_prices.get(self.opt.co2_storage_resource_id or 'CO2_STORAGE', {}).get(t, 0.0)
                                tr_price = self.data.time_series.resource_prices.get(self.opt.co2_transport_resource_id or 'CO2_TRANSPORT', {}).get(t, 0.0)
                                yr_ccs_st += (s_price + tr_price) * captured_tons_per_unit * act_v
            ccs_st_costs.append(yr_ccs_st / 1_000_000.0)

        # 1. Create Resources-Only Figure
        df_res_export = df_res_costs.copy()
        df_res_export['Year'] = years
        fig_res = build_opex_figure(
            df_opex=df_res_export,
            years=years,
            theme="report",
            title=f"RESOURCES OPEX: {self.scenario_name}"
        )
        self._save_plotly_figure(fig_res, "Resources_Opex", show_png=show_png)

        # 2. Create Total OPEX Figure (including tech, tax, etc.)
        df_bars = pd.concat([df_res_costs, df_tech_costs], axis=1)
        if sum(ccs_st_costs) > 1e-4: df_bars['CCS STORAGE & TRANSPORT'] = ccs_st_costs
        if sum(dac_opex) > 1e-4: df_bars['DAC OPEX'] = dac_opex
        df_bars['CARBON TAX (GROSS)'] = tax_costs
        if sum(credit_costs) > 1e-4: df_bars['CARBON CREDITS'] = credit_costs
            
        df_bars = df_bars.loc[:, (df_bars.abs() > 1e-4).any(axis=0)]
        df_total_export = df_bars.copy()
        df_total_export['Year'] = years
        df_total_export['TOTAL ANNUAL OPEX (M€)'] = df_bars.sum(axis=1).values
        
        fig_total = build_opex_figure(
            df_opex=df_total_export,
            years=years,
            theme="report",
            title=f"TOTAL ANNUAL OPEX: {self.scenario_name}"
        )
        
        self._save_plotly_figure(fig_total, "Total_Annual_Opex", show_png=show_png)
        self.charts_data.append(("Generalized Operational Expenditure Breakdown", df_total_export))

    def _plot_carbon_tax_and_avoided(self, df: pd.DataFrame, show_png: bool = True):
        """Generates an interactive Plotly visualization using the high-fidelity builder."""
        df_plot = df.copy()
        if 'Year' in df_plot.columns:
            years = df_plot['Year'].tolist()
        else:
            years = list(self.opt.years)
        
        standard_tax = df_plot['Standard_Tax_Cost_MEuros'].tolist() if 'Standard_Tax_Cost_MEuros' in df_plot.columns else [0.0]*len(years)
        indirect_tax = df_plot['Indirect_Tax_Cost_MEuros'].tolist() if 'Indirect_Tax_Cost_MEuros' in df_plot.columns else [0.0]*len(years)
        penalty_costs = df_plot['Penalty_Cost_MEuros'].tolist() if 'Penalty_Cost_MEuros' in df_plot.columns else [0.0]*len(years)
        avoided_reduced = (df_plot['Really_Avoided_CO2_kt'] * 1000 * df_plot['Tax_Price'] / 1_000_000.0).tolist() if 'Really_Avoided_CO2_kt' in df_plot.columns else [0.0]*len(years)
        avoided_captured = (df_plot['Captured_CO2_kt'] * 1000 * df_plot['Tax_Price'] / 1_000_000.0).tolist()
        ccfd_refunds = df_plot['CCfD_Refund_MEuros'].tolist() if 'CCfD_Refund_MEuros' in df_plot.columns else None

        fig = build_carbon_tax_figure(
            years=years,
            standard_tax=standard_tax,
            indirect_tax=indirect_tax,
            penalties=penalty_costs,
            avoided_reduced=avoided_reduced,
            avoided_captured=avoided_captured,
            ccfd_refunds=ccfd_refunds,
            theme="report",
            title=f"CARBON TAX & PENALTIES: {self.scenario_name}"
        )

        self._save_plotly_figure(fig, "Carbon_Tax", show_png=show_png)
        self.charts_data.append(("Carbon Tax and Penalties Balance", df_plot))

    def _plot_carbon_prices(self, show_png: bool = True):
        """Generates a high-fidelity carbon price visualization."""
        years = self.years
        carbon_prices = [self.data.time_series.carbon_prices.get(y, 0.0) for y in years]
        penalties = [self.data.time_series.carbon_penalties.get(y, 0.0) for y in years]
        effective_prices = [p * (1.0 + x) for p, x in zip(carbon_prices, penalties)]

        strike_prices = []
        if hasattr(self.opt, "ccfd_used_vars"):
            for (t_inv, p_id, t_id), var in self.opt.ccfd_used_vars.items():
                if var.varValue and var.varValue > 0.5:
                    tech = self.data.technologies[t_id]
                    ccfd_p = self.data.ccfd_params
                    base_p = self.data.time_series.carbon_prices.get(t_inv, 0.0)
                    strike_val = (1.0 + ccfd_p.eua_price_pct) * base_p
                    start_yr = t_inv + tech.implementation_time
                    end_yr = start_yr + ccfd_p.duration
                    contract_years = [y for y in years if start_yr <= y < end_yr]
                    if contract_years:
                        strike_prices.append({'name': tech.name or t_id, 'val': strike_val, 'years': contract_years})

        fig = build_carbon_price_figure(
            years=years,
            market_prices=carbon_prices,
            effective_prices=effective_prices,
            strike_prices=strike_prices,
            theme="report",
            title=f"CARBON PRICE & POLICY: {self.scenario_name}"
        )

        self._save_plotly_figure(fig, "Carbon_Prices", show_png=show_png)
        
        simplified_strikes = []
        for y in years:
            active_at_y = [s['val'] for s in strike_prices if y in s['years']]
            simplified_strikes.append(max(active_at_y) if active_at_y else 0.0)

        df_cp = pd.DataFrame({
            'Year': years,
            'Market_Price': carbon_prices,
            'Effective_Price': effective_prices,
            'Penalty_Factor': penalties,
            'Strike_Price': simplified_strikes
        })
        if 'Year' in df_cp.columns:
            df_cp.set_index('Year', inplace=True)
        self.charts_data.append(("Carbon Price & Policy Trajectory", df_cp))

    def _plot_transition_costs(self, df_costs: pd.DataFrame, df_finance: pd.DataFrame, df_emis: pd.DataFrame, show_png: bool = True):
        """Plots the stacked cumulative costs and savings of the ecological transition using Plotly."""
        df_costs = df_costs.copy()
        if 'Year' in df_costs.columns:
            df_costs.set_index('Year', inplace=True)

        def get_robust_price(r_id, yr):
            """Get resource price with exact lookup first, then fuzzy fallback."""
            resource_prices = self.data.time_series.resource_prices
            exact_price = float(resource_prices.get(r_id, {}).get(yr, 0.0) or 0.0)
            if exact_price != 0.0:
                return exact_price

            rid_base = r_id[3:] if isinstance(r_id, str) and r_id.startswith('EN_') else r_id
            for k, yearly_prices in resource_prices.items():
                if not isinstance(k, str) or not isinstance(yearly_prices, dict):
                    continue
                k_base = k[3:] if k.startswith('EN_') else k
                if rid_base == k_base or rid_base in k_base or k_base in rid_base:
                    fuzzy_price = float(yearly_prices.get(yr, 0.0) or 0.0)
                    if fuzzy_price != 0.0:
                        return fuzzy_price
            return 0.0
            
        # 1. Baseline Calculation
        baseline_data = []
        b_res_costs = {} # (t, res_id) -> cost
        primary_emis = self._primary_emission_resource()
        for t in self.years:
            yr_idx = list(self.years).index(t)
            b_emissions = 0.0
            b_consumptions = {res_id: 0.0 for res_id in self.data.resources}
            
            for p_id, process in self.opt.entity.processes.items():
                up_rate = 0.0
                if 'UP' in process.valid_technologies:
                    up_tech = self.data.technologies['UP']
                    up_imp = up_tech.impacts.get('ALL') or up_tech.impacts.get(primary_emis)
                    if up_imp and (up_imp['type'] == 'variation' or up_imp['type'] == 'up'):
                        up_rate = abs(up_imp['value'])
                
                up_factor = (1.0 - up_rate) ** yr_idx
                p_base_emis = (self.opt.entity.process_emission_baseline(process, primary_emis)) * up_factor
                b_emissions += p_base_emis
                for res_id in self.data.resources:
                    if not self._is_primary_emission_resource(res_id):
                        p_base_cons = (self.opt.entity.base_consumptions.get(res_id, 0.0) * process.consumption_shares.get(res_id, 0.0)) * up_factor
                        b_consumptions[res_id] += p_base_cons
            
            allocated_emis_share = sum(p.emission_shares.get(primary_emis, 0.0) for p in self.opt.entity.processes.values())
            b_emissions += self.opt.entity.base_emissions * (1.0 - allocated_emis_share)
            for res_id in self.data.resources:
                if not self._is_primary_emission_resource(res_id):
                    allocated_share = sum(p.consumption_shares.get(res_id, 0.0) for p in self.opt.entity.processes.values())
                    b_consumptions[res_id] += self.opt.entity.base_consumptions.get(res_id, 0.0) * (1.0 - allocated_share)
            
            tax_price = self.data.time_series.carbon_prices.get(t, 0.0)
            if self.opt.entity.sv_act_mode == "PI":
                fq_pct = self.data.time_series.carbon_quotas_pi.get(t, 0.0)
            else:
                fq_pct = self.data.time_series.carbon_quotas_norm.get(t, 0.0)
            
            taxed_co2_b = b_emissions * (1.0 - fq_pct) if fq_pct <= 1.0 else max(0.0, b_emissions - fq_pct)
            b_tax_cost = taxed_co2_b * tax_price / 1_000_000.0
            
            for res_id, cons_val in b_consumptions.items():
                if res_id == 'CO2_EM': continue
                price = get_robust_price(res_id, t)
                if price > 0:
                    b_res_costs[(t, res_id)] = (cons_val * price) / 1_000_000.0
            
            baseline_data.append({'Year': t, 'Baseline_Tax': b_tax_cost})
        
        df_b = pd.DataFrame(baseline_data).set_index('Year')
        years = list(df_b.index)
        df_annual = pd.DataFrame(index=years)
        
        # --- POSITIVE COSTS (Efforts) ---
        df_fin = df_finance.set_index('Year')
        df_annual['Self-funded CAPEX'] = df_fin['Out_of_Pocket_CAPEX (M€)'] + df_fin['Principal_Repayment (M€)']
        df_annual['Bank Loan Service'] = df_fin['Interest_Paid (M€)']
        
        tech_opex = []
        for t in years:
            yr_opex = 0.0
            for p_id, process in self.opt.entity.processes.items():
                for t_id in process.valid_technologies:
                    tech = self.data.technologies[t_id]
                    if tech.is_continuous_improvement:
                        continue
                    act_v = self._get_safe_var_value(self.opt.active_vars[(t, p_id, t_id)])
                    if act_v > 0.01:
                        cap_opex = 1.0
                        if tech.opex_per_unit:
                            if tech.opex_unit == 'tCO2': cap_opex = self.opt.entity.process_emission_baseline(process, primary_emis)
                            elif 'MW' in str(tech.opex_unit).upper():
                                best_val = self._process_primary_energy_consumption(process)
                                cap_opex = best_val / self.opt.entity.annual_operating_hours
                        current_opex = tech.opex_by_year.get(t, tech.opex)
                        yr_opex += (current_opex * cap_opex / process.nb_units) * act_v
            if self.data.dac_params.active:
                dac_total_v = self._get_safe_var_value(self.opt.dac_total_capacity_vars.get(t))
                yr_opex += dac_total_v * self.data.dac_params.opex_by_year.get(t, 0.0)
            tech_opex.append(yr_opex / 1_000_000.0)
        df_annual['Tech & DAC OPEX'] = tech_opex
        
        cred_costs = []
        for t in years:
            cred_v = self._get_safe_var_value(self.opt.credit_purchased_vars.get(t))
            cred_costs.append((cred_v * self.data.credit_params.cost_by_year.get(t, 0.0)) / 1_000_000.0)
        df_annual['Voluntary Carbon Credits'] = cred_costs
        
        # --- NEGATIVE COSTS (Savings/Aids) ---
        public_aids = []
        for t in years:
            grant_total = sum(self._get_safe_var_value(self.opt.grant_amt_vars.get((t, p_id, t_id))) 
                             for p_id, proc in self.opt.entity.processes.items() 
                             for t_id in proc.valid_technologies 
                             if (t, p_id, t_id) in self.opt.grant_amt_vars and self.opt.grant_amt_vars.get((t, p_id, t_id)) is not None)
            ccfd_refund = df_emis.set_index('Year').at[t, 'CCfD_Refund_MEuros']
            public_aids.append(-(grant_total / 1_000_000.0 + ccfd_refund))
        df_annual['Public Aids (Grants & CCfD)'] = public_aids
        
        # Carbon Tax: Show Actual Scenario Tax vs Baseline Offset
        actual_tax = df_emis.set_index('Year')['Tax_Cost_MEuros']
        df_annual['Carbon Tax (Actual)'] = actual_tax
        df_annual['Baseline Carbon Tax Offset'] = -df_b['Baseline_Tax'] # Benefit of avoiding BAU tax
        
        # Per-resource delta (Added vs Suppressed)
        actual_res_costs = {}
        for t in years:
            for res_id in self.data.resources:
                if not self._is_primary_emission_resource(res_id):
                    cons_val = self.opt.cons_vars.get((t, res_id))
                    cons_val_numeric = self._get_safe_var_value(cons_val)
                    price = get_robust_price(res_id, t)
                    if price > 0:
                        actual_res_costs[(t, res_id)] = (cons_val_numeric * price) / 1_000_000.0
        
        add_res_costs = []
        avoid_res_savings = []
        for t in years:
            yr_add = 0.0
            yr_avoid = 0.0
            for res_id in self.data.resources:
                if not self._is_primary_emission_resource(res_id):
                    b_c = b_res_costs.get((t, res_id), 0.0)
                    a_c = actual_res_costs.get((t, res_id), 0.0)
                    delta = a_c - b_c
                    if delta > 1e-4: yr_add += delta
                    elif delta < -1e-4: yr_avoid += delta
            add_res_costs.append(yr_add)
            avoid_res_savings.append(yr_avoid)

        df_annual['Additional Resource Cost'] = add_res_costs
        df_annual['Avoided Resource Saving'] = avoid_res_savings
        
        pos_cols = ['Self-funded CAPEX', 'Bank Loan Service', 'Tech & DAC OPEX', 'Voluntary Carbon Credits', 'Additional Resource Cost', 'Carbon Tax (Actual)']
        neg_cols = ['Public Aids (Grants & CCfD)', 'Baseline Carbon Tax Offset', 'Avoided Resource Saving']
        pos_cols = [c for c in pos_cols if c in df_annual.columns and df_annual[c].abs().sum() > 1e-3]
        neg_cols = [c for c in neg_cols if c in df_annual.columns and df_annual[c].abs().sum() > 1e-3]

        # Prefix columns for strict categorization parity
        for c in pos_cols:
            if c in df_annual.columns:
                df_annual = df_annual.rename(columns={c: f"Effort: {c}"})
        for c in neg_cols:
            if c in df_annual.columns:
                df_annual = df_annual.rename(columns={c: f"Saving: {c}"})
        
        # Update column lists with prefixes
        pos_cols_prefixed = [f"Effort: {c}" for c in pos_cols if f"Effort: {c}" in df_annual.columns]
        neg_cols_prefixed = [f"Saving: {c}" for c in neg_cols if f"Saving: {c}" in df_annual.columns]

        # Plotly Construction (Using Shared Module)
        fig = build_transition_cost_figure(
            df_annual=df_annual,
            years=years,
            pos_cols=pos_cols_prefixed,
            neg_cols=neg_cols_prefixed,
            investment_cap=self.data.reporting_toggles.investment_cap,
            title=f"ECOLOGICAL TRANSITION: {self.scenario_name}"
        )

        # Finalize and Save
        self._save_plotly_figure(fig, "Transition_Cost", show_png=show_png)
        
        # Data for Excel
        df_store = df_annual[pos_cols_prefixed + neg_cols_prefixed].copy()
        df_store.index.name = "Year"
        df_store = df_store.reset_index()
        self.charts_data.append(("TRANSITION_COST_HIGH_FIDELITY", df_store))
        
        self.charts_data.append(("Ecological Transition Cumulative Balance", df_store.set_index('Year').cumsum().reset_index()))

    def _plot_interest_paid(self, df_finance: pd.DataFrame, show_png: bool = True):
        df_plot = df_finance.copy()
        if 'Year' in df_plot.columns:
            df_plot.set_index('Year', inplace=True)
            
        if 'Interest_Paid (M€)' not in df_plot.columns or df_plot['Interest_Paid (M€)'].sum() < 1e-4:
            return

        fig = build_interest_paid_figure(
            df_plot=df_plot,
            years=list(df_plot.index),
            theme="report",
            title=f"BANK LOANS & INTEREST: {self.scenario_name}"
        )
        
        self._save_plotly_figure(fig, "Interest_Paid", show_png=show_png)
        self.charts_data.append(("Bank Loans: Interest Paid", df_plot[['Interest_Paid (M€)']]))

    def _plot_prices(self, show_png: bool = True):
        """Plots all prices used in the simulation using the high-fidelity builder."""
        price_series = {}
        if self.data.time_series.carbon_prices:
            price_series['EUA'] = {
                'data': self.data.time_series.carbon_prices,
                'name': 'EUA (Carbon Price)',
                'unit': '€/tCO2',
                'color': '#2C3E50'
            }
            
        colors = ['#2980B9', '#8E44AD', '#D35400', '#C0392B', '#16A085', '#273746', '#F39C12', '#BDC3C7']
        color_idx = 0
        for r_id, p_dict in self.data.time_series.resource_prices.items():
            if p_dict:
                res = self.data.resources.get(r_id)
                name = res.name if res and res.name else r_id
                unit = res.unit if res and res.unit else 'unit'
                price_series[r_id] = {
                    'data': p_dict,
                    'name': name,
                    'unit': f"€/{unit}",
                    'color': colors[color_idx % len(colors)]
                }
                color_idx += 1
                
        if not price_series:
            return

        fig = build_simulation_prices_figure(
            price_series=price_series,
            years=list(self.years),
            theme="report",
            title=f"PRICE PARAMETERS: {self.scenario_name}"
        )
        
        self._save_plotly_figure(fig, "Simulation_Prices", show_png=show_png)
        self._save_plotly_figure(fig, "Data_Used", show_png=show_png)
        
        df_prices = pd.DataFrame({info['name']: info['data'] for info in price_series.values()}).sort_index()
        df_prices.index.name = 'Year'
        self.charts_data.append(("Simulation Prices", df_prices))

    def _plot_co2_abatement_cost(self, df: pd.DataFrame, show_png: bool = True):
        """Plot the MAC curve using the high-fidelity builder."""
        if df is None or df.empty:
            self._save_plotly_placeholder(
                "CO2_Abatement",
                "MARGINAL ABATEMENT COST",
                "No valid abatement projects found to build a MAC curve.",
                show_png=show_png
            )
            return
        df_plot = df.sort_values(by='MAC (€/tCO2)')
        
        avg_co2_price = sum(self.data.time_series.carbon_prices.values()) / len(self.data.time_series.carbon_prices) if self.data.time_series.carbon_prices else 0.0
        total_abated = df_plot['Total Abated (tCO2)'].sum()

        fig = build_mac_figure(
            df_plot=df_plot,
            avg_carbon_price=avg_co2_price,
            total_simulation_abatement=total_abated,
            theme="report",
            title=f"CO2 ABATEMENT COST (MAC): {self.scenario_name}"
        )

        self._save_plotly_figure(fig, "CO2_Abatement", show_png=show_png)
        self.charts_data.append(("CO2 Abatement Cost (MAC)", df_plot))

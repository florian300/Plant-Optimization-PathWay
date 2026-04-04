import openpyxl

file_path = 'PathFinder input.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

found = False
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if 'BROWNIEN' in cell.value.upper():
                    print(f"FOUND in openpyxl: '{cell.value}' in sheet '{sheet_name}', cell {cell.coordinate}")
                    found = True

if not found:
    print("openpyxl search also failed to find 'BROWNIEN' in cells. This is very strange.")
